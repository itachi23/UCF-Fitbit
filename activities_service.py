from static.urls import ACTIVITIES_URL
from datetime import datetime
import requests
import pandas as pd
import openpyxl
import sys
import os 
current = os.path.dirname(os.path.realpath(__file__))
parent = os.path.dirname(current)
sys.path.append(parent)
import utils

def get_activities_data(url, access_token):
    headers = {"authorization":f"Bearer {access_token}"}
    response = requests.get(url,headers=headers)
    print(response.status_code)
    if(response.status_code != 200):
        print(response.text)
        return None
    data = response.json()["activities"]
    return data

def get_url(user_id,from_date):
    return ACTIVITIES_URL.replace("{}",user_id).replace("[]",from_date)

def is_response_valid(response,prev_log_id):
    if(len(response) == 0):
        return False
    result = response is not None and prev_log_id != response[len(response) - 1]["logId"]
    if(result == False):
        return False
    return True

def get_columns(data_frame):
    return list(data_frame.columns.values)

def add_data_to_sheet(sheet, data,file,workbook):
    df = pd.DataFrame(data)
    sheet.append(get_columns(df))
    for row_data in df.itertuples(index = False):
        sheet.append(row_data)
    workbook.save(file)

def parse_date(date):
    parsed_date = datetime.strptime(date, "%Y-%m-%dT%H:%M:%S.%f%z")
    formatted_date = parsed_date.strftime("%Y-%m-%d")
    return formatted_date

def get_new_file_name(date):
    s = date.split('_')
    new_s = str(int(s[1]) + 1)
    return s[0]+"_"+new_s

def create_file_name(new_date,prev_date,prev_file_name):
    new_date = parse_date(new_date)
    if(new_date == prev_date):
        file_name  = get_new_file_name(prev_file_name)
    else:
        file_name = new_date+"_0"
    return file_name

prev_file = ""
prev_log_id = ""
token_workbook = utils.get_workbook("tokens.xlsx",[])
sheet = token_workbook.active
root_directory = r"D:\fitbit_user_data"
exclusion_list = {"caloriesLink","detailsLink","logId","manualValuesSpecified","tcxLink","heartRateLink","hasActiveZoneMinutes"}

data_type = {"activeZoneMinutes":("Active Zone Minutes","minutesInHeartRateZones"),"activityLevel":("Activity Level","activityLevel"),
             "heartRateZones":("Heartrate Zones","heartRateZones"),"source":("Source","source")}

for row in sheet.iter_rows(min_row = 2,values_only = True):
    from_date = "2020-01-01"
    id = utils.get_user_id(row[1])
    print(f"processing user {id}")
    user_path = os.path.join(root_directory,id)
    user_path = os.path.join(user_path,"Excercise")
    url = get_url(row[4],from_date)
    response = get_activities_data(url,row[3])
    if(response is None):
        continue
    while(is_response_valid(response,prev_log_id)):
        file_name = ""
        prev_date = ""
        prev_file_name = ""
        for data in response:
            workbook = openpyxl.Workbook()
            row_data = [{}]
            first_sheet = workbook.active
            file_name = prev_file_name = create_file_name(data["startTime"],prev_date,prev_file_name)
            file_name = file_name + ".xlsx"
            final_path = os.path.join(user_path,file_name)
            for key in data:
                if(key in exclusion_list):
                    continue
                elif(key in data_type):
                    d = None
                    if(key == "activeZoneMinutes"):
                        d = data[key]["minutesInHeartRateZones"]
                    else:
                        d = data[key]
                    sheet = workbook.create_sheet(title = data_type[key][0])
                    add_data_to_sheet(sheet, d, final_path, workbook)
                else:
                    row_data[0][key] = data[key]
            add_data_to_sheet(workbook[workbook.sheetnames[0]], row_data,final_path,workbook)
            prev_date = parse_date(data["startTime"])
            prev_log_id = data["logId"]
        url = get_url(row[4],prev_date)
        response = get_activities_data(url,row[3])        
                    