from openpyxl import Workbook
from openpyxl import load_workbook 
import os

def create_workbook(column_names,file_name):
    workbook = Workbook()
    sheet = workbook.active
    for col_index, column_name in enumerate(column_names, start=1):
        sheet.cell(row=1, column=col_index, value=column_name)
    workbook.save(file_name)
    return workbook

def get_workbook(file_name,column_names):
    workbook = ""
    if(file_exists(file_name)):
        workbook = load_workbook(get_file_path(file_name))
    else:
        workbook = create_workbook(column_names,file_name)
    
    return workbook

def get_file_path(file_name):
    current_directory = os.path.dirname(__file__)
    excel_file_path = os.path.join(current_directory, file_name)
    return excel_file_path

def file_exists(file_name):
    return False if(not os.path.exists(get_file_path(file_name))) else True

def get_user_id(user_mail):
    id = ""
    for i in range(len(user_mail)):
        value = user_mail[i]
        if value.isdigit():
            id = id + value
    return id