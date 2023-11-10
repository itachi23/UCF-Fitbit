import os
import sys
from requests.exceptions import RequestException, Timeout, HTTPError, ConnectionError
current = os.path.dirname(os.path.realpath(__file__))
parent = os.path.dirname(current)
sys.path.append(parent)
from static.urls import SLEEP_URL
from services.static.paths import SLEEP_LOG_PATH, ROOT_DIR
from dateutil.relativedelta import relativedelta
from datetime import date,datetime
from loguru import logger
import logconfig
import requests
import pandas as pd
import openpyxl 
import utils


def add_data_to_sheet(sheet, data):
    df = pd.DataFrame(data)
    sheet.append(get_columns(df))
    for row_data in df.itertuples(index = False):
        sheet.append(row_data)

def get_columns(data_frame):
    return list(data_frame.columns.values)

def get_sleep_data(url, access_token):
    response = make_http_request(url, access_token)
    if(response.status_code != 200):
        logger.error(response.text)
        return None
    data = response.json()
    if("sleep" not in data):
        logger.error(response.text)
        return None
    return data["sleep"]

def make_http_request(url,access_token):
    headers = {"authorization":f"Bearer {access_token}","accept-language":"en_US"}

    try:
        response = requests.get(url,headers=headers)
        return response
    
    except ConnectionError as conn_err:
        logger.error(conn_err)
        return None

    except Timeout as timeout_err:
        logger.error(timeout_err)
        return None
    
    except HTTPError as http_err:
        logger.error(http_err)
        return None
    
    except RequestException as req_err:
        logger.error(req_err)
        return None

def get_url(user_id,from_date):
    return SLEEP_URL.replace("{}",user_id).replace("[]",from_date)

def construct_summary_data(summary):
    summary_item = []
    for key in summary:
        element = {}
        element["sleep_stage"] = key
        t = tuple(summary[key].items())
        for item in t:
            element[item[0]] = item[1]
        summary_item.append(element)
    return summary_item

def get_new_file_name(date):
    s = date.split('.')[0].split('_')
    new_s = str(int(s[1]) + 1)
    return s[0]+"_"+new_s+".xlsx"

def is_given_date_equals_today(given_date):
    given_date = datetime.strptime(given_date,"%Y-%m-%d")
    result = given_date == date.today()
    print(f"g date {given_date}")
    print(date.today())
    return given_date == date.today()

def get_from_date(path):
    files = os.listdir(path)
    if(len(files) == 0):
        return -1
    files = [f for f in files if os.path.isfile(os.path.join(path, f))]
    files.sort(key=lambda x: os.path.getmtime(os.path.join(path, x)), reverse=True)
    l = files[0].split('_')[0].split('-')
    last_date = date(int(l[1]), int(l[2]), int(l[3]))
    from_date = last_date + relativedelta(days=1)
    from_date = from_date.strftime("%Y-%m-%d")
    return from_date


def write_data_to_file(response,prev_date,user_path,id):
    for data in response:
        new_date = data["dateOfSleep"]

        if(is_given_date_equals_today(new_date)):
            break

        workbook = openpyxl.Workbook()
        if(new_date == prev_date):
            file_name = get_new_file_name(prev_file)
        else:
            file_name = id + "-" + new_date +"_0.xlsx"
        prev_date = new_date
        prev_file = file_name
        sheet1 = workbook.active
        final = os.path.join(user_path,file_name)
        item = [{"startTime":data["startTime"], "endTime":data["endTime"],"efficiency":data["efficiency"], "isMainSleep":data["isMainSleep"],
                "MinutesAsleep":data["minutesAsleep"],"minutesAwake":data["minutesAwake"],"timeInBed":data["timeInBed"]}]       
        add_data_to_sheet(sheet1, item)

        sheet2 = workbook.create_sheet(title="Details")
        item = data["levels"]["data"]
        add_data_to_sheet(sheet2, item)

        sheet3 = workbook.create_sheet(title="Summary")
        item = construct_summary_data(data["levels"]["summary"])
        add_data_to_sheet(sheet3,item)
        workbook.save(final)     

def is_response_valid(response,prev_log_id):
    if response is None:
        return False
    if(len(response) == 0):
        return False
    result =  prev_log_id != response[len(response) - 1]["logId"]
    if(result == False):
        return False
    return True

def configure_path(mail_id):
    id = utils.get_user_id(mail_id)
    user_path = os.path.join(ROOT_DIR,id)
    user_path = os.path.join(user_path,"Sleep")
    return [id, user_path]


prev_file = ""
prev_log_id = ""
token_workbook = utils.get_workbook("tokens.xlsx",[])
sheet = token_workbook.active
logger.configure(**logconfig.configure_logs(SLEEP_LOG_PATH))

for row in sheet.iter_rows(min_row = 5,max_row = 5,values_only = True):
    from_date = "2020-01-01"
    id, user_path = configure_path(row[1])
    logger.info(f"processing user {id}")
    
    value = get_from_date(user_path)
    from_date = value if value != -1 else from_date

    url = get_url(row[4],from_date)
    response = get_sleep_data(url,row[3])
    if(response is None):
        logger.error(f"error while getting sleep data for user {id}")
        continue
    while(is_response_valid(response,prev_log_id)):
        prev_date = ""
        write_data_to_file(response,prev_date,user_path,id)
        from_date = response[len(response) - 1]["dateOfSleep"]
        prev_log_id = response[len(response) - 1]["logId"]
        url = get_url(row[4],from_date)
        response = get_sleep_data(url,row[3])
    logger.info(f"process complete for user {id}")
        
    








