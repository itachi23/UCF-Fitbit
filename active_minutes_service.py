from requests.exceptions import RequestException, Timeout, HTTPError, ConnectionError
from dateutil.relativedelta import relativedelta
from datetime import datetime
from static.urls import ACTIVE_MINUTES_URL
import requests
from openpyxl import Workbook,load_workbook
import pandas as pd
import sys
import os 
current = os.path.dirname(os.path.realpath(__file__))
parent = os.path.dirname(current)
sys.path.append(parent)
import utils
from enum import Enum

class ACTIVE_MINUTES_TYPE(Enum):
    SEDENTARY = ("minutesSedentary", "activities-minutesSedentary")
    LIGHTLY = ("minutesLightlyActive", "activities-minutesLightlyActive")
    FAIRLY = ("minutesFairlyActive", "activities-minutesFairlyActive")
    VERY_ACTIVE = ("minutesVeryActive", "activities-minutesVeryActive")

def create_active_minutes_workbook(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Date", "sedentary_minutes","light_active_minutes","fairly_active_minutes","very_active_minutes"])
    workbook.save(path)
    return workbook

def modify_url(active_minutes_type, from_date, user_id, url):
    return url.replace("()", active_minutes_type).replace("[]", from_date).replace("{}",user_id)

def find_start_date(url,access_token, data_type):
    respone = make_http_request(url, access_token)
    if respone is None:
        return -1
    for data in respone[data_type]:
         if (data["value"] != 1440):
            return data["dateTime"]
         
    
def is_file_empty(path):
    Workbook = load_workbook(path)
    sheet = Workbook.worksheets[0]
    row = sheet[2]
    for i in range(2):
        if row[i].value is None:
            return True
    return False

def get_from_date(workbook):
    sheet = workbook.worksheets[0]
    return sheet.cell(row=sheet.max_row, column=1).value

def get_data(url, access_token, data_type):
    response = make_http_request(url, access_token)
    if response is None:
        return None
    return response[data_type]

def write_data_to_file(workbook, path, collection, id): 
    data = []
    row_len = len(collection)
    column_len = len(collection[0])
    for i in range(column_len):
        d = dict()
        date = collection[0][i]["dateTime"]
        d["dateTime"] = date
        for j in range(row_len):          
            d[j] = collection[j][i]["value"]
        data.append(d)
    add_data_to_sheet(workbook.active, data)

    try:
        workbook.save(path)
        return f"data saved successfully for user {id}"
    except Exception as e:
        print(e)
        return f"error while saving the data for user {id}"
    
def add_data_to_sheet(sheet, data):
    df = pd.DataFrame(data)
    for row_data in df.itertuples(index=False):
        sheet.append(row_data)

def make_http_request(url,access_token):
    headers = {"authorization":f"Bearer {access_token}","accept-language":"en_US"}

    try:
        response = requests.get(url,headers=headers)
        return response.json()
    
    except ConnectionError as conn_err:
        print(conn_err)
        return None

    except Timeout as timeout_err:
        print(timeout_err)
        return None
    
    except HTTPError as http_err:
        print(http_err)
        return None
    
    except RequestException as req_err:
        print(req_err)
        return None

root_directory = r"D:\fitbit_user_data"
token_workbook = utils.get_workbook("tokens.xlsx",[])
token_sheet = token_workbook.active

for row in token_sheet.iter_rows(min_row = 89,values_only = True):
    from_date = "2023-05-01"
    id = utils.get_user_id(row[1])
    url = ""
    active_minutes_workbook = ""
    collection = []
    print(f"processing user {id}")

    user_path = os.path.join(root_directory,id)
    user_path = os.path.join(user_path,"Physical_Activity\\Active_Minutes")
    final_path = os.path.join(user_path,f"active_minutes_{id}.xlsx")

    if not os.path.exists(final_path) or (os.path.exists(final_path) and is_file_empty(final_path)):
        if(os.path.exists(final_path)):
            os.remove(final_path)
        active_minutes_workbook = create_active_minutes_workbook(final_path)
        url = modify_url(ACTIVE_MINUTES_TYPE.SEDENTARY.value[0],from_date, row[4], ACTIVE_MINUTES_URL)
        from_date = find_start_date(url, row[3], ACTIVE_MINUTES_TYPE.SEDENTARY.value[1])

    else:
        active_minutes_workbook = load_workbook(final_path)
        from_date = datetime.strptime(get_from_date(active_minutes_workbook),'%Y-%m-%d') + relativedelta(days = 1)
        from_date = from_date.strftime('%Y-%m-%d')

    if from_date == -1:
        continue
    
    for active_minutes_type in ACTIVE_MINUTES_TYPE:
        url = modify_url(active_minutes_type.value[0],from_date, row[4], ACTIVE_MINUTES_URL)
        data = get_data(url, row[3], active_minutes_type.value[1])
        if(data is not None):
            collection.append(data)

    if(len(collection) < 3):
        continue
    
    result = write_data_to_file(active_minutes_workbook, final_path, collection, id)
    print(result)
