import os
import sys
current = os.path.dirname(os.path.realpath(__file__))
parent = os.path.dirname(current)
sys.path.append(parent)
import utils
import requests
from static.paths import ROOT_DIR
from openpyxl import load_workbook, Workbook
from datetime import datetime
from dateutil.relativedelta import relativedelta
from requests.exceptions import RequestException, Timeout, HTTPError, ConnectionError

def get_token_workbook():
    return utils.get_workbook("tokens.xlsx",[])

def configure_path(mail_id, folder_name):
    id = utils.get_user_id(mail_id)
    user_path = os.path.join(ROOT_DIR,id)
    user_path = os.path.join(user_path,folder_name)
    return [id, user_path]

def is_file_empty(path):
    Workbook = load_workbook(path)
    sheet = Workbook.worksheets[0]
    row = sheet[2]
    for i in range(2):
        if row[i].value is None:
            return True
    return False

def create_workbook(path, column_names):  
    workbook =  Workbook()
    sheet = workbook.active
    sheet.append(column_names)
    workbook.save(path)
    return workbook

def get_from_date(workbook):
    sheet = workbook.worksheets[0]
    date =  sheet.cell(row=sheet.max_row, column=1).value
    date =  datetime.strptime(date,'%Y-%m-%d') + relativedelta(days = 1)
    return date.strftime('%Y-%m-%d')

def convert_to_date(date_str):
    return datetime.strptime(date_str,"%Y-%m-%d")

def is_response_valid(response):
    if not isinstance(response, requests.Response):
        return False

    return True

def get_last_sync_date(url,access_token):
    response = make_http_request(url,access_token)  
    
    if not is_response_valid(response):
        return None
    
    data = response.json()

    if(len(data) == 0):
        
        return data

    last_sync = None
    for devices in data:
        if (devices.get("deviceVersion") == "Charge 5"):
            last_sync = devices.get("lastSyncTime")
            break
        
    return last_sync
    

def make_http_request(url,access_token):
    headers = {"authorization":f"Bearer {access_token}","accept-language":"en_US"}

    try:
        response = requests.get(url,headers=headers)
        response.raise_for_status()
        return response
    
    except ConnectionError as conn_err:
        return conn_err

    except Timeout as timeout_err:
        return timeout_err
    
    except HTTPError as http_err:
        return http_err
    
    except RequestException as req_err:
        return req_err