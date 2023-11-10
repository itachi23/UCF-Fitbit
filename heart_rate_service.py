from static.urls import HEART_RATE_URL
from datetime import datetime
from dateutil.relativedelta import relativedelta
import requests
from openpyxl import Workbook,load_workbook
import pandas as pd
import openpyxl
import sys
import os 
current = os.path.dirname(os.path.realpath(__file__))
parent = os.path.dirname(current)
sys.path.append(parent)
import utils


def get_url(user_id,from_date):
    return HEART_RATE_URL.replace("{}",user_id).replace("[]",from_date)

def get_from_date(workbook):
    sheet = workbook.worksheets[1]
    return sheet.cell(row=sheet.max_row, column=1).value

def initialize_data(url,from_date):
    start_date = datetime.strptime(from_date, '%Y-%m-%d')
    end_date = start_date + relativedelta(years = 1)
    start_date = start_date.strftime('%Y-%m-%d')
    end_date_str = end_date.strftime('%Y-%m-%d')
    print(end_date_str)
    url = url.replace("today",end_date_str)
    print(url)
    return [start_date,end_date_str,url]

def find_start_date(url,access_token,from_date):

    result = None
    new_start_date = ""
    flag = False
    start_date,end_date_str,url = initialize_data(url,from_date)
    
    while datetime.strptime(start_date,'%Y-%m-%d') < datetime.now():
        response = make_http_request(url,access_token)
        if(response.status_code != 200):
            print(response.text)
            return None
        data = response.json()
        for item in data["activities-heart"]:
            if("caloriesOut" in item["value"]["heartRateZones"][0]):
                result = item["dateTime"]
                flag = True
                break
            new_start_date = item["dateTime"]
        
        if(flag):
            break
    
        new_start_date = datetime.strptime(new_start_date, '%Y-%m-%d')
        new_end_date = new_start_date + relativedelta(years = 1)
        new_end_date_str = new_end_date.strftime('%Y-%m-%d')
        url = url.replace(end_date_str,new_end_date_str).replace(start_date,end_date_str)
        start_date = end_date_str
        end_date_str = new_end_date 
        # print(url)
    return result
    

def create_heart_rate_workbook(user_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Resting heart rate"
    sheet.append(["Date", "Resting Heart Rate"])
    hrz_sheet = workbook.create_sheet("Heart Rate Zones")
    c_hrz_sheet = workbook.create_sheet("Custom Heart Rate Zones")
    column_names = ["Date","Calories Out","max heart rate","min heart rate","minutes in zone","Zone name"]
    hrz_sheet.append(column_names)
    c_hrz_sheet.append(column_names)
    workbook.save(os.path.join(user_path,"heart_rate.xlsx"))
    return workbook


def add_data_to_sheet(sheet,data,date):
    df = pd.DataFrame(data)
    df_as_list = df.values.tolist()
    for row_data in df_as_list:
        sheet.append([date] + row_data)
    
      
def write_heart_rate_to_file(url,access_token,path,workbook):
    response = make_http_request(url,access_token)

    if(response.status_code != 200):
        print(response.text)
        return None
    data = response.json()

    for d in data["activities-heart"]:
        date = d["dateTime"]
        value = d["value"]
        n = 2
        for item in value:
            if(item == "restingHeartRate"):
                add_data_to_sheet(workbook.worksheets[n],[value[item]],date)
            else:
                add_data_to_sheet(workbook.worksheets[n],value[item],date)
            n-=1
        workbook.save(path)
    return 1
            

def make_http_request(url,access_token):
    headers = {"authorization":f"Bearer {access_token}"}
    response = requests.get(url,headers=headers)
    return response


def is_file_empty(path):
    Workbook = load_workbook(path)
    sheet = Workbook.worksheets[1]
    row = sheet[2]
    for i in range(6):
        print(f"value is {row[i].value}")
        if row[i].value is None:
            return True
    return False


root_directory = r"D:\fitbit_user_data"
token_workbook = utils.get_workbook("tokens.xlsx",[])
token_sheet = token_workbook.active

for row in token_sheet.iter_rows(min_row = 9,values_only = True):
    heart_rate_work_book = ""
    from_date = "2023-01-01"
    id = utils.get_user_id(row[1])
    url = ""
    print(f"processing user {id}")

    user_path = os.path.join(root_directory,id)
    user_path = os.path.join(user_path,"Heart")
    final_path = os.path.join(user_path,"heart_rate.xlsx")


    if not os.path.exists(final_path) or (os.path.exists(final_path) and is_file_empty(final_path)):
        if(os.path.exists(final_path)):
            os.remove(final_path)
        heart_rate_work_book = create_heart_rate_workbook(user_path)
        url = get_url(row[4],from_date)
        from_date = find_start_date(url,row[3],from_date)
        print(from_date)

    else:
       heart_rate_work_book = load_workbook(final_path)
       from_date = get_from_date(heart_rate_work_book)

    if(from_date is None):
        continue
    url = get_url(row[4],from_date)
    result = write_heart_rate_to_file(url,row[3],final_path,heart_rate_work_book)
    if(result is None):
        print(f"There was a problem while getting heart rate data for user {id}")
    else:
        print(f"data retreived successfully for {id}")