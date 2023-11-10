import os
import sys
from requests.exceptions import RequestException, Timeout, HTTPError, ConnectionError
current = os.path.dirname(os.path.realpath(__file__))
parent = os.path.dirname(current)
sys.path.append(parent)
from static.urls import CALORIES_URL,ACIVITY_CALORIES_URL
from services.static.paths import CALORIES_PATH, ROOT_DIR
from dateutil.relativedelta import relativedelta
from datetime import date,datetime
from loguru import logger
import logconfig
import requests
import pandas as pd
from openpyxl import Workbook, load_workbook
import utils
from enum import Enum

class CALORIES_TYPE(Enum):
    CALORIES = (CALORIES_URL, "activities-calories") 
    ACIVITY_CALORIES_URL = (ACIVITY_CALORIES_URL,"activities-activityCalories" )


def construct_final_data(collection):
    final_data = []
    column_len = len(collection[0])
    for i in range(column_len):
        d = dict()
        d["date"] = collection[0][i]["dateTime"]
        d["calories"] = collection[0][i]["value"]
        d["activity_calories"] = collection[1][i]["value"]
        final_data.append(d)
    return final_data

def write_data_to_file(collection, workbook, path):
    data = construct_final_data(collection)
    # print(data)   
    add_data_to_sheet(workbook.active,data)
    try:
        workbook.save(path)
        logger.info(f"data saved successfully for user {id}")
    except Exception as e:
        logger.error(e)

def add_data_to_sheet(sheet,data):
    df = pd.DataFrame(data)
    for row_data in df.itertuples(index = False):
        sheet.append(row_data)

def get_calories_data(url, access_token, type):
    response = make_http_request(url, access_token)
    data = response.json()
    if response is None:
        return None
    
    if type not in data:
        logger.error(f"{response.status_code}  {response.text}")
        return None

    return data[type]

def make_http_request(url,access_token):
    headers = {"authorization":f"Bearer {access_token}","accept-language":"en_US"}

    try:
        response = requests.get(url,headers=headers)
        return response
    
    except ConnectionError as conn_err:
        logger.error(conn_err)
        return None

    except Timeout as timeout_err:
        logger.error(timeout_err)
        return None
    
    except HTTPError as http_err:
        logger.error(http_err)
        return None
    
    except RequestException as req_err:
        logger.error(req_err)
        return None
    
def modify_url(url,user_id,from_date):
    yesterday = date.today() - relativedelta(days = 1)
    yesterday = yesterday.strftime('%Y-%m-%d')
    return url.replace("{}",user_id).replace("[]",from_date).replace("today", yesterday)

def initialize_from_date(workbook):
    last_date = get_last_date(workbook)
    new_date = datetime.strptime(last_date, '%Y-%m-%d') + relativedelta(days = 1)
    new_date = new_date.strftime('%Y-%m-%d')
    return new_date

def get_last_date(workbook):
    sheet = workbook.worksheets[0]
    return sheet.cell(row=sheet.max_row, column=1).value

def create_calories_workbook(path):
    workbook =  Workbook()
    sheet = workbook.active
    sheet.append(["Date", "Caliroes", "Activity_Calories"])
    workbook.save(path)
    return workbook

def configure_path(mail_id):
    id = utils.get_user_id(mail_id)
    user_path = os.path.join(ROOT_DIR,id)
    user_path = os.path.join(user_path,"Calories")
    final_path = os.path.join(user_path,f"calories_{id}.xlsx")
    return [id, final_path]

def is_file_empty(path):
    Workbook = load_workbook(path)
    sheet = Workbook.worksheets[0]
    row = sheet[2]
    for i in range(2):
        if row[i].value is None:
            return True
    return False

token_workbook = utils.get_workbook("tokens.xlsx",[])
token_sheet = token_workbook.active
calories_workbook = ""

logger.configure(**logconfig.configure_logs(CALORIES_PATH))

for row in token_sheet.iter_rows(min_row = 4,values_only = True):
    
    collection = []
    from_date = "2023-05-01"
    id, user_path = configure_path(row[1])
    logger.info(f"processing user {id}")
    file_exists = os.path.exists(user_path)

    if not file_exists:
        calories_workbook = create_calories_workbook(user_path)

    else:
        if is_file_empty(user_path):
            os.remove(user_path)
            calories_workbook = create_calories_workbook(user_path)

        else:
            calories_workbook = load_workbook(user_path)
            from_date = initialize_from_date(calories_workbook)
   
    for calories_type in CALORIES_TYPE:
        url = modify_url(calories_type.value[0], row[4], from_date)
        data = get_calories_data(url, row[3], calories_type.value[1])
        if(data is None):
            logger.error(f"Failed to get {calories_type.value[1]}")
            break
        collection.append(data)

    if(len(collection) < 2):
        logger.error(" Failed to collect required data ")
        continue

    write_data_to_file(collection, calories_workbook, user_path)