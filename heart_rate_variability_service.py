from static.urls import HEART_RATE_VARIABILITY_URL
from static.paths import  HEART_RATE_VARIABILITY_PATH
import requests
from dateutil.relativedelta import relativedelta
from datetime import datetime
from openpyxl import Workbook,load_workbook
import pandas as pd
import sys
import os 
from loguru import logger

current = os.path.dirname(os.path.realpath(__file__))
parent = os.path.dirname(current)
sys.path.append(parent)
import utils
import logconfig
import service_utils as su

def modify_url(url, user_id, from_date):
    return url.replace("{}",user_id).replace("[]",from_date)

def get_next_dates(end_date_str):
    new_start_date = datetime.strptime(end_date_str, '%Y-%m-%d') + relativedelta(days = 1)
    new_end_date = new_start_date + relativedelta(days = 29)
    new_start_date_str = new_start_date.strftime('%Y-%m-%d')
    new_end_date_str = new_end_date.strftime('%Y-%m-%d')
    return [new_start_date_str, new_end_date_str]

def write_data_to_file(data,workbook,path, id):
    add_data_to_sheet(workbook.active,data)
    try:
        workbook.save(path)
        logger.info(f"data saved for user {id}")
    except Exception as e:
        print(e)
        logger.error(f"Error while writing data to file {e}")


def add_data_to_sheet(sheet,data):
    df = pd.DataFrame(data)
    for row_data in df.itertuples(index = False):
        sheet.append(row_data)

def iterate(data):
    collection = []
    for item in data:
        d = dict()
        d["dateTime"] = item["dateTime"]
        d["dailyRmssd"] = item["value"]["dailyRmssd"]
        d["deepRmssd"] = item["value"]["deepRmssd"]
        collection.append(d)
    return collection

def is_response_valid(response):
    if not isinstance(response, requests.Response):
        logger.error(response)
        return False
    if "hrv" in response.json():
        return True
    return False

def collect_data(url, from_data, access_token):
    final_data = []
    last_date = datetime.now() - relativedelta(days = 1)
    start_date = from_date
    end_date = datetime.strptime(from_data,'%Y-%m-%d') + relativedelta(days = 29)
    end_date_str = end_date.strftime('%Y-%m-%d')
    last_date_str = last_date.strftime('%Y-%m-%d')

    if end_date > datetime.now():
        url = url.replace('today', last_date_str)
        response = su.make_http_request(url, access_token)
        if not is_response_valid(response):
            return final_data
        
        return iterate(response.json()['hrv'])
    
    else:
        url = url.replace('today', end_date_str)
        while datetime.strptime(start_date,'%Y-%m-%d') < last_date:
            response = su.make_http_request(url, access_token)
            if not is_response_valid(response):
                return final_data
            data = response.json()
            final_data = final_data + iterate(data['hrv'])

            [new_start_date_str,new_end_date_str] = get_next_dates(end_date_str)
            new_end_date = datetime.strptime(new_end_date_str,'%Y-%m-%d')
            if(new_end_date > datetime.now()):
                url = url.replace(start_date,new_start_date_str).replace(end_date_str,last_date_str)
            else:
                url = url.replace(end_date_str,new_end_date_str).replace(start_date,new_start_date_str)
            start_date = new_start_date_str
            end_date_str = new_end_date_str

    return final_data
            

sheet = su.get_token_workbook().active
column_names = ["Date", "dailyRmssd", "deepRmssd"]
logger.configure(**logconfig.configure_logs(HEART_RATE_VARIABILITY_PATH))

for row in sheet.iter_rows(min_row = 84,max_row = 85,values_only = True):
    from_date = "2023-05-01"
    id, user_path = su.configure_path(row[1], "Heartrate Variability")
    user_path = os.path.join(user_path,f"hrv_{id}.xlsx")
    logger.info(f"processing user {id}")

    file_exists = os.path.exists(user_path)

    if not file_exists:
        calories_workbook = su.create_workbook(user_path, column_names)

    else:
        if su.is_file_empty(user_path):
            os.remove(user_path)
            calories_workbook = su.create_workbook(user_path, column_names)

        else:
            calories_workbook = load_workbook(user_path)
            from_date = su.get_from_date(calories_workbook)

    url = modify_url(HEART_RATE_VARIABILITY_URL, row[4], from_date)
    data = collect_data(url, from_date, row[3])
    write_data_to_file(data, calories_workbook, user_path, id)