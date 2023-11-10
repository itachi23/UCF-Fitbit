from static.urls import HEAR_RATE_INTRADAY_URL, DEVICES_URL
from static.paths import  HEART_RATE_INTRADAY_LOG_PATH
import requests
from dateutil.relativedelta import relativedelta
from datetime import date,datetime
from openpyxl import Workbook,load_workbook
import pandas as pd
import sys
import os 
from loguru import logger
import logconfig
import service_utils as su



def create_file(date, path, id):
    workbook = Workbook()
    columns = ["Time", "Value"]
    sheet = workbook.active
    sheet.append(columns)
    file = os.path.join(path,f"{date}_{id}.xlsx")
    workbook.save(file)
    return workbook

def write_data_to_file(workbook,data, path):
    sheet = workbook.active
    try:
        add_data_to_sheet(sheet,data)
        workbook.save(path)
        logger.info(f"data saved for user {id}")
    except Exception as e:
        logger.error(f"Error while writing data to file {e}")

def add_data_to_sheet(sheet,data):
    df = pd.DataFrame(data)
    for row_data in df.itertuples(index = False):
        sheet.append(row_data)

def get_intraday_data(access_token, intraday_url, devices_url, from_date, id, path):
    intraday_data = []
    response = su.get_last_sync_date(devices_url, access_token)
    if(response is None):
        logger.error(response)
    
    date = su.convert_to_date(response)

    if(date == from_date):
        return 
        
    date = date - relativedelta(days = 1)

    while from_date != date:
        from_date_str = datetime.strftime(from_date, "%Y-%m-%d")
        response = su.make_http_request(intraday_url, access_token)

        if not su.is_response_valid(response):
            return intraday_data
        
        data = response.json()

        if "activities-heart-intraday" in data:
            return intraday_data
        
        data = data["activities-heart-intraday"]["dataset"]

        if len(data) == 0:
            return intraday_data
        
        workbook = create_file(from_date_str, path, id)

        write_data_to_file(workbook, data, path)
        from_date = from_date + relativedelta(days = 1)
        

def get_from_date(path):
    files = os.listdir(path)
    if(len(files) == 0):
        return -1
    files = [f for f in files if os.path.isfile(os.path.join(path, f))]
    files.sort(key=lambda x: os.path.getmtime(os.path.join(path, x)), reverse=True)
    l = files[0].split('_')[0].split('-')
    last_date = date(int(l[0]), int(l[1]), int(l[2]))
    from_date = last_date + relativedelta(days=1)
    from_date = from_date.strftime("%Y-%m-%d")
    return from_date

sheet = su.get_token_workbook().active

logger.configure(**logconfig.configure_logs(HEART_RATE_INTRADAY_LOG_PATH))

for row in sheet.iter_rows(min_row = 84,max_row = 85,values_only = True):
    from_date = "2023-05-01"
    id, user_path = su.configure_path(row[1], "\\Intraday\\Heart_Rate")
    logger.info(f"processing user {id}")

    value = get_from_date(user_path)
    from_date = value if value != -1 else from_date
    get_intraday_data(row[3], HEAR_RATE_INTRADAY_URL, DEVICES_URL, from_date, id, user_path)