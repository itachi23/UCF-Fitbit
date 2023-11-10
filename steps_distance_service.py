from static.urls import STEPS_URL,DISTANCE_URL
import requests
from openpyxl import Workbook,load_workbook
from requests.exceptions import RequestException, Timeout, HTTPError, ConnectionError
import pandas as pd
import sys
import os 
current = os.path.dirname(os.path.realpath(__file__))
parent = os.path.dirname(current)
sys.path.append(parent)
import utils
from enum import Enum

class DATATYPE(Enum):
    STEPS = "activities-steps"
    DISTANCE = "activities-distance"

def create_steps_distance_workbook(user_path,id):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Steps"
    sheet.append(["Date", "Value"])
    distance_sheet = workbook.create_sheet("Distance")
    distance_sheet.append(["Date", "Miles"])
    workbook.save(os.path.join(user_path,f"steps_distance_{id}.xlsx"))
    return workbook

def is_file_empty(path):
    Workbook = load_workbook(path)
    sheet = Workbook.worksheets[0]
    row = sheet[2]
    for i in range(2):
        if row[i].value is None:
            return True
    return False


def get_url(user_id,url_type):
    return url_type.replace("{}",user_id).replace("[]",from_date)


def write_data_to_file(data,workbook,sheet,path,type):
    add_data_to_sheet(sheet,data[type])
    try:
        workbook.save(path)
        return f"Successfully saved {type}"
    except Exception as e:
        print(e)
        return f"Error while writing {type} data to file"


def add_data_to_sheet(sheet,data):
    df = pd.DataFrame(data)
    for row_data in df.itertuples(index = False):
        sheet.append(row_data)


def perform(url_type,user_id,access_token,path,workbook,sheet,type):
    url = get_url(user_id,url_type)
    response = make_http_request(url,access_token)
    if response is None:
        return "Error while making HTTP request"
    return write_data_to_file(response,workbook,sheet,path,type)



def get_from_date(workbook):
    sheet = workbook.worksheets[1]
    return sheet.cell(row=sheet.max_row, column=1).value
    

def make_http_request(url,access_token):
    headers = {"authorization":f"Bearer {access_token}","accept-language":"en_US"}

    try:
        response = requests.get(url,headers=headers)
        return response.json()
    
    except ConnectionError as conn_err:
        print(conn_err)
        return None

    except Timeout as timeout_err:
        print(timeout_err)
        return None
    
    except HTTPError as http_err:
        print(http_err)
        return None
    
    except RequestException as req_err:
        print(req_err)
        return None
    
    
    

root_directory = r"D:\fitbit_user_data"
token_workbook = utils.get_workbook("tokens.xlsx",[])
token_sheet = token_workbook.active
from_date = "2023-05-01"

for row in token_sheet.iter_rows(min_row = 2,values_only = True):
    
    id = utils.get_user_id(row[1])
    url = ""
    steps_distance_workbook = ""
    print(f"processing user {id}")

    user_path = os.path.join(root_directory,id)
    user_path = os.path.join(user_path,"Physical_Activity\\Steps")
    final_path = os.path.join(user_path,f"steps_distance_{id}.xlsx")
    print(final_path)

    if not os.path.exists(final_path) or (os.path.exists(final_path) and is_file_empty(final_path)):
        if(os.path.exists(final_path)):
            os.remove(final_path)
        steps_distance_workbook = create_steps_distance_workbook(user_path,id)
    else:
        steps_distance_workbook = load_workbook(final_path)
        from_date = get_from_date(steps_distance_workbook)

    result = perform(STEPS_URL,row[4],row[3],final_path,steps_distance_workbook, 
            steps_distance_workbook.worksheets[0], DATATYPE.STEPS.value)
    print(result)

    result = perform(DISTANCE_URL,row[4],row[3],final_path,steps_distance_workbook, 
            steps_distance_workbook.worksheets[1], DATATYPE.DISTANCE.value)
    print(result)
    