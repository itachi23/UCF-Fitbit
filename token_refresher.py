from requests.exceptions import RequestException, Timeout, HTTPError, ConnectionError
from services.static.paths import SYNC_STATUS_PATH
from loguru import logger
from openpyxl import Workbook
from openpyxl import load_workbook 
import requests
from datetime import datetime

def is_access_token_expired(created_date):
    created_date = datetime.strptime(created_date, "%Y-%m-%d %H:%M:%S")
    difference = (datetime.now() - created_date).total_seconds()/3600
    logger.info(f"difference = {8 - difference}")
    return True if(8 - difference < 1) else False

def generate_new_access_token(TOKEN_URL,refresh_token,user_id,headers,token_workbook,row_number):
    try:
        body = f"grant_type=refresh_token&refresh_token={refresh_token}"
        check_if_file_is_already_open(user_id, row_number, token_workbook)
        response = make_http_request(TOKEN_URL, body, headers)
        if response is None:
            return None
        data = response.json()
        refresh_token = data['refresh_token']
        access_token = data['access_token']
        created_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        update_workbook(access_token,refresh_token,token_workbook,row_number,created_date)
        return access_token
    except Exception as e:
        logger.error(f"An error occurred while generating new access token {e}")
        if(response is not None):
            logger.error(response.text)
        return None
    
def make_http_request(url, body, headers):
    try:
        response = requests.post(url,data = body,headers = headers)
        return response
    
    except ConnectionError as conn_err:
        logger.error(conn_err)
        return None

    except Timeout as timeout_err:
        logger.error(timeout_err)
        return None
    
    except HTTPError as http_err:
        logger.error(http_err)
        return None
    
    except RequestException as req_err:
        logger.error(req_err)
        return None

def update_workbook(access_token,refresh_token,token_workbook, row_number,created_date):
    sheet = token_workbook.active
    refresh_token_index = 3
    access_token_index = 4
    created_date_index = 6
    sheet.cell(row=row_number, column= refresh_token_index, value=refresh_token) 
    sheet.cell(row=row_number, column= access_token_index, value=access_token)
    sheet.cell(row=row_number, column= created_date_index, value=created_date)
    logger.info("updating workbook")
    token_workbook.save("tokens.xlsx")

def check_if_file_is_already_open(user_id, row_number, token_workbook):
    user_id_index = 5
    sheet = token_workbook.active
    sheet.cell(row=row_number, column= user_id_index, value=user_id)
    token_workbook.save("tokens.xlsx")
