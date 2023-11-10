from static.urls import VO2MAX_URL
import requests
from dateutil.relativedelta import relativedelta
from datetime import datetime
from openpyxl import Workbook,load_workbook
from requests.exceptions import RequestException, Timeout, HTTPError, ConnectionError
import pandas as pd
import sys
import os 
current = os.path.dirname(os.path.realpath(__file__))
parent = os.path.dirname(current)
sys.path.append(parent)
import utils


def modify_url(url,user_id,from_date):
    return url.replace("{}",user_id).replace("[]",from_date)

def get_from_date(workbook):
    sheet = workbook.worksheets[0]
    return sheet.cell(row=sheet.max_row, column=1).value

def create_vo2_max_workbook(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Date", "Value"])
    workbook.save(path)
    return workbook

def is_file_empty(path):
    Workbook = load_workbook(path)
    sheet = Workbook.worksheets[0]
    row = sheet[2]
    for i in range(2):
        if row[i].value is None:
            return True
    return False

def make_http_request(url,access_token):
    headers = {"authorization":f"Bearer {access_token}","accept-language":"en_US"}

    try:
        response = requests.get(url,headers=headers)
        return response.json()
    
    except ConnectionError as conn_err:
        print(conn_err)
        return None

    except Timeout as timeout_err:
        print(timeout_err)
        return None
    
    except HTTPError as http_err:
        print(http_err)
        return None
    
    except RequestException as req_err:
        print(req_err)
        return None
    
def write_data_to_file(data,workbook,path):
    add_data_to_sheet(workbook.active,data)
    try:
        workbook.save(path)
        return f"Successfully saved"
    except Exception as e:
        print(e)
        return f"Error while writing data to file"


def add_data_to_sheet(sheet,data):
    df = pd.DataFrame(data)
    for row_data in df.itertuples(index = False):
        sheet.append(row_data)
    
def get_next_dates(end_date_str):
    new_start_date = datetime.strptime(end_date_str, '%Y-%m-%d') + relativedelta(days = 1)
    new_end_date = new_start_date + relativedelta(days = 29)
    new_start_date_str = new_start_date.strftime('%Y-%m-%d')
    new_end_date_str = new_end_date.strftime('%Y-%m-%d')
    return [new_start_date_str, new_end_date_str]

def initialize_data(url,from_date):
    start_date = datetime.strptime(from_date, '%Y-%m-%d')
    end_date = start_date + relativedelta(days = 29)
    start_date = start_date.strftime('%Y-%m-%d')
    end_date_str = end_date.strftime('%Y-%m-%d')
    print(end_date_str)
    url = url.replace("today",end_date_str)
    print(url)
    return [start_date,end_date_str,url]

def iterate(data):
    collection = []
    for item in data:
        d = dict()
        d["dateTime"] = item["dateTime"]
        d["Vo2_max"] = item["value"]["vo2Max"]
        collection.append(d)
    return collection

def construct_data(url, from_data, access_token):
    end_date = datetime.strptime(from_data,'%Y-%m-%d') + relativedelta(days = 29)
    end_date_str = end_date.strftime('%Y-%m-%d')
    start_date = from_date
    data = []

    if(end_date < datetime.now()):
        url = url.replace('today', end_date_str)

    else:
        response = make_http_request(url, access_token)
        if(response is not None):
            return iterate(response["cardioScore"])
        return data

    while datetime.strptime(start_date,'%Y-%m-%d') < datetime.now():
        response = make_http_request(url, access_token)
        if(response is None):
            return data

        data = data + iterate(response["cardioScore"])

        [new_start_date_str,new_end_date_str] = get_next_dates(end_date_str)
        new_end_date = datetime.strptime(new_end_date_str,'%Y-%m-%d')
        if(new_end_date > datetime.now()):
            url = url.replace(start_date,new_start_date_str).replace(end_date_str,'today')
        else:
            url = url.replace(end_date_str,new_end_date_str).replace(start_date,new_start_date_str)
        # print(url)
        start_date = new_start_date_str
        end_date_str = new_end_date_str

    return data

def find_start_date(url,from_date,access_token):
    result = -1
    start_date,end_date_str,url = initialize_data(url,from_date)
    while datetime.strptime(start_date,'%Y-%m-%d') < datetime.now():
        response = make_http_request(url,access_token)

        if response is not None:
            if("cardioScore" in response):
                data = response["cardioScore"]
                if(len(data) > 0):
                    result = data[0]["dateTime"]
                    return result
        
        [new_start_date_str,new_end_date_str] = get_next_dates(end_date_str)
        new_end_date = datetime.strptime(new_end_date_str,'%Y-%m-%d')     
        if(new_end_date > datetime.now()):
            url = url.replace(start_date,new_start_date_str).replace(end_date_str,'today')
        else:
            url = url.replace(end_date_str,new_end_date_str).replace(start_date,new_start_date_str)
        # print(url)

        start_date = new_start_date_str
        end_date_str = new_end_date_str
    return result

root_directory = r"D:\fitbit_user_data"
token_workbook = utils.get_workbook("tokens.xlsx",[])
token_sheet = token_workbook.active

for row in token_sheet.iter_rows(min_row = 75,values_only = True):

    from_date = "2023-05-01"
    id = utils.get_user_id(row[1])
    url = ""
    vo2_max_workbook = ""
    print(f"processing user {id}")

    user_path = os.path.join(root_directory,id)
    user_path = os.path.join(user_path,"VO2_Max")
    final_path = os.path.join(user_path,f"vo2_max_{id}.xlsx")

    if not os.path.exists(final_path) or (os.path.exists(final_path) and is_file_empty(final_path)):
        if(os.path.exists(final_path)):
            os.remove(final_path)
        
        vo2_max_workbook = create_vo2_max_workbook(final_path)
        n_url = modify_url(VO2MAX_URL,row[4],from_date)
        from_date = find_start_date(n_url,from_date,row[3])

    else:
        vo2_max_workbook = load_workbook(final_path)
        from_date = datetime.strptime(get_from_date(vo2_max_workbook),'%Y-%m-%d') + relativedelta(days = 1)
        from_date = from_date.strftime('%Y-%m-%d')
        

    if(from_date == -1):
        print("data is empty")
        continue


    url = modify_url(VO2MAX_URL,row[4],from_date)
    data = construct_data(url, from_date, row[3])
    result = write_data_to_file(data,vo2_max_workbook,final_path) if len(data) > 0 else "data is empty"
    print(result)